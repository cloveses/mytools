import xlrd
import openpyxl
import xlsxwriter

def get_xls_datas(filename, row_deal_function=None, tail_rows=0, start_row=1):
    """start_row＝1 有一行标题行；gred_end=1 末尾行不导入"""
    """row_del_function 为每行的数据类型处理函数，不传则对数据类型不作处理 """
    wb = xlrd.open_workbook(filename)
    ws = wb.sheets()[0]
    nrows = ws.nrows
    datas = []
    for i in range(start_row, nrows - tail_rows):
        row = ws.row_values(i)
        # print(row)
        if row_deal_function:
            row = row_deal_function(row)
        datas.append(row)
    return datas

def get_xls_sheets_data(filename, row_deal_function=None, tail_rows=0, start_row=1, retain_headline=False):
    """ start_row＝1 有一行标题行；gred_end=1 末尾行不导入
        row_del_function 为每行的数据类型处理函数，不传则对数据类型不作处理 
        retain_headline 保留首个sheet的标题行标志"""

    # names = data.sheet_names()
    # table = data.sheet_by_name(sheet_name)

    wb = xlrd.open_workbook(filename)
    # ws = wb.sheets()[0]
    names = wb.sheet_names()
    nsheets = wb.nsheets
    datas = []
    for index, sheet_index in enumerate(range(nsheets)):
        ws = wb.sheet_by_index(sheet_index)
        nrows = ws.nrows
        if retain_headline and index == 0:
            start = 0
        else:
            start = start_row
        for i in range(start,nrows-tail_rows):
            row = ws.row_values(i)
            if row_deal_function:
                row = row_deal_function(row)
            datas.append(row)
    return datas


def get_xlsx_datas(filename, row_deal_function=None, tail_rows=0, start_row=0, first_sheet=True, retain_headline=True):
    '''
    first_sheet 是否只取第一张表的数据
    '''
    wb = openpyxl.load_workbook(filename, read_only=True)
    datas = []
    for index, sheet in enumerate(wb.worksheets):
        if index == 0 and retain_headline:
            min_row = 0
        else:
            min_row = start_row
        for row in sheet.iter_rows(min_row=min_row, values_only=True):
            if row_deal_function:
                datas.append(row_deal_function(row))
            else:
                datas.append(row)
        if first_sheet:
            break
    return datas


def get_files(directory):
    files = []
    files = os.listdir(directory)
    files = [f for f in files if f.endswith('.xls') or f.endswith('.xlsx')]
    files = [os.path.join(directory,f) for f in files]
    return files
    
def countit(datas):
    # 对(datas)列表中的数据项进行分类统计并输出
    from collections import Counter
    c = Counter(datas)
    for k,v in c.items():
        print(k,v,sep='\t')
    return c

def save_datas_xlsx(filename,datas):
    #将一张表的信息写入电子表格中XLSX文件格式
    w = xlsxwriter.Workbook(filename)
    w_sheet = w.add_worksheet('sheet1')
    for rowi,row in enumerate(datas):
        for coli,celld in enumerate(row):
            w_sheet.write(rowi,coli,celld)
    w.close()

def summary_col(filename,col_seq_num,res_filename='sum.xlsx'):
    # 统计指定电子表格文件中的指定序号列到指定的文件中
    # filename 指定电子表格文件
    # col_seq_num 列序号从0开始
    # res_filename 统计结果存放文件名
    cols_data = get_data_cols(filename)
    cols_data = cols_data[col_seq_num]
    res = countit(cols_data)
    res = [[k,v] for k,v in res.items()]
    save_datas_xlsx(res_filename,res)

def merge_files_data(mydir, res_filename, row_deal_function=None, headline_rows=0, tail_rows=0):
    # 合并指定目录(mydir)下的分表数据到一个电子表格文件(res_filename)中的一张表中
    if not os.path.exists(mydir):
        print('Directory is not exist.')
        return
    filenames = get_files(mydir)
    datass = []
    for index, filename in enumerate(filenames):
        if filename.endswith('xls'):
            if index == 0:
                datas = get_xls_sheets_data(filename, row_deal_function=row_deal_function, tail_rows=tail_rows, start_rows=headline_rows, retain_headline=True)
            else:
                datas = get_xls_sheets_data(filename, row_deal_function=row_deal_function, tail_rows=tail_rows, start_rows=headline_rows)
        else:
            if index == 0:
                datas = get_xlsx_datas(filename, row_deal_function=row_deal_function, tail_rows=tail_rows, start_rows=headline_rows, retain_headline=True, first_sheet=False)
            else:
                datas = get_xlsx_datas(filename, row_deal_function=row_deal_function, tail_rows=tail_rows, start_rows=headline_rows, first_sheet=False)
        datass.extend(datas)

    save_datas_xlsx(res_filename,datass)

if __name__ == '__main__':
    print(get_xlsx_datas('示范高中指标分解.xlsx'))